# -*- coding: utf-8 -*-
"""
Created on Sat May 23 16:31:43 2020

@author: Damon
"""

import tweepy
import time
import os
from openpyxl import load_workbook

os.system('cls')


def read_last_seen(file1):
    file_read = open(file1,'r')
    last_seen_id = int(file_read.read().strip())
    file_read.close()
    return last_seen_id


def store_last_seen(file1,last1):
    file_write = open(file1,'w')
    file_write.write(str(last1))
    file_write.close()
    return


def reply_to_mentions():
    return_val = 0
    last_id = read_last_seen(LAST_SEEN_FILE)
    #print('\nLast ID:',last_id,'\n')
    mentions = api.mentions_timeline(last_id,tweet_mode='extended')
    number_of_mentions = len(mentions)
    print('\nReceived ' + str(number_of_mentions) + ' new tweets\n')
    if number_of_mentions > 0:
        hash_tweets = sum(1 for tweet1 in mentions if HASHTAG1 in 
                          tweet1.full_text.upper())
        print('\n Replying only to',hash_tweets,'tweets with hashtag',HASHTAG1,':\n')
        for tweet in reversed(mentions):
            #if '#PYBOT_TEST' in tweet.text.upper():
            if HASHTAG1 in tweet.full_text.upper() and \
                tweet.user.screen_name != 'LightpathData':
                print('Replying to ID:',tweet.id)
                #print(tweet.user.screen_name + ':' + tweet.user.name + ' ' + tweet.full_text)
                #print('\n----\n')
                api.update_status('Hey @'+tweet.user.screen_name + ' #PyBot_Test worked (v7.1).',tweet.id)
                api.create_favorite(tweet.id)
                #print('Reply Sent\n----\n')
            if 'PYBOT_END' in tweet.full_text.upper():
                return_val+=1
        print('Saving', str(tweet.id),'as last_id')
        store_last_seen(LAST_SEEN_FILE,tweet.id)
    return return_val

KEY_FILE = 'TwitterKeys.xlsx'

LAST_SEEN_FILE = 'Last_Seen.txt'
HASHTAG1 = '#PYBOT_TEST'

ind1 = 0

wb = load_workbook(filename = KEY_FILE,read_only=True)
ws = wb['TwitterAuth']

consumer_key = ws['B1'].value
consumer_secret = ws['B2'].value

key = ws['B3'].value
secret = ws['B4'].value


auth = tweepy.OAuthHandler(consumer_key, consumer_secret)
auth.set_access_token(key, secret)


api = tweepy.API(auth)


while ind1 == 0:
    ind1 = reply_to_mentions()
#      time.sleep(15)  #For spot-checks
    time.sleep(7200)   # 2 Hours

